"""
Discord Activity Tracker Bot
============================

An open-source Discord bot that tracks user activity in both text and voice channels
and stores data in an Excel workbook (`user_activity.xlsx`).

Features:
- Tracks message activity and voice joins
- Provides user-specific and basic server-wide activity summaries
- Built using discord.py and openpyxl

Author: Your Name
License: MIT
Repository: https://github.com/Fordywhat/discord-activity-tracker
"""

import os
from dotenv import load_dotenv
import discord
from discord.ext import commands
from discord import app_commands
import openpyxl
import datetime

# ---------------------------------------------------------------------------
# Loading Enviornment Variables
# ---------------------------------------------------------------------------

print('------')
print('Loading environment variables...')

load_dotenv()

DISCORD_TOKEN = os.getenv("DISCORD_TOKEN")
GUILD_ID = int(os.getenv("GUILD_ID"))  # Convert to int for command registration


# ---------------------------------------------------------------------------
# Workbook Setup
# ---------------------------------------------------------------------------

print('------')
print('Loading workbook...')

''' Replace user-activity.xlsx in WORKBOOK_PATH to change workbook location '''
WORKBOOK_PATH = "user_activity.xlsx" 
workbook = openpyxl.load_workbook(WORKBOOK_PATH)
user_worksheet = workbook.active

HEADER_ROW = 1
HEADER_DATA_ROW = 2
LOG_HEADER_ROW = 3
LOG_DATA_ROW = 4

USER_ID = 1
EVENT_TYPE = 1
INVITED_BY = 2
EVENT_TIME = 2
INVITE_DATE = 3
EVENT_CONTENT = 3
TOTAL_MESSAGE = 4
TOTAL_CALL = 5

# ---------------------------------------------------------------------------
# Helper Functions
# ---------------------------------------------------------------------------
def createUserSheet(user_id, invited_by, invite_date):
    """
    Create a new sheet for a user in the workbook.

    Args:
    user_id (str): The ID of the user.
    invited_by (str): The ID of the user who invited them.
    invite_date (str): The date the user was invited.
    """
    print('------')
    print(f'Creating user sheet for {user_id}...')

    for id in workbook.sheetnames:
        if id == user_id:
            print(f'User sheet for {user_id} already exists. Skipping creation.')
            return

    user_worksheet = workbook.create_sheet(title=user_id)

    row_1 = ['USER_ID', 'INVITED_BY', 'INVITE_DATE', 'TOTAL_MESSAGES', 'TOTAL_CALLS']
    row_2 = [user_id, invited_by, invite_date, 0, 0]
    row_3 = ['EVENT_TYPE', 'EVENT_TIME', 'EVENT_CONTENT']

    user_worksheet.cell(row=HEADER_ROW, column=USER_ID, value=row_1[0])
    user_worksheet.cell(row=HEADER_ROW, column=INVITED_BY, value=row_1[1])
    user_worksheet.cell(row=HEADER_ROW, column=INVITE_DATE, value=row_1[2])
    user_worksheet.cell(row=HEADER_ROW, column=TOTAL_MESSAGE, value=row_1[3])
    user_worksheet.cell(row=HEADER_ROW, column=TOTAL_CALL, value=row_1[4])

    user_worksheet.cell(row=HEADER_DATA_ROW, column=USER_ID, value=row_2[0])
    user_worksheet.cell(row=HEADER_DATA_ROW, column=INVITED_BY, value=row_2[1])
    user_worksheet.cell(row=HEADER_DATA_ROW, column=INVITE_DATE, value=row_2[2])
    user_worksheet.cell(row=HEADER_DATA_ROW, column=TOTAL_MESSAGE, value=row_2[3])
    user_worksheet.cell(row=HEADER_DATA_ROW, column=TOTAL_CALL, value=row_2[4]) 

    user_worksheet.cell(row=LOG_HEADER_ROW, column=EVENT_TYPE, value=row_3[0])
    user_worksheet.cell(row=LOG_HEADER_ROW, column=EVENT_TIME, value=row_3[1])
    user_worksheet.cell(row=LOG_HEADER_ROW, column=EVENT_CONTENT, value=row_3[2])

    print(f'User sheet for {user_id} created successfully.')

    print('Saving workbook...')

    workbook.save(WORKBOOK_PATH)

def updateSpreadsheet(user_id, type, time, content):
    """
    Update the activity spreadsheet for a user.

    Args:
    user_id (str): The ID of the user.
    event_type (str): The type of event ('invite', 'message', 'voice').
    event_time (str): The timestamp of the event.
    event_content (str): The content of the event (e.g., message text).
    """
    print('------')
    print(f'Updating spreadsheet for user {user_id} with event {type} at {time}.')

    for id in workbook.sheetnames:
        if id == user_id:
            user_worksheet = workbook[id]

            user_worksheet.insert_rows(LOG_DATA_ROW)

            # Update the new row with the event data
            user_worksheet.cell(row=LOG_DATA_ROW, column=EVENT_TYPE, value=type)
            user_worksheet.cell(row=LOG_DATA_ROW, column=EVENT_TIME, value=time)
            user_worksheet.cell(row=LOG_DATA_ROW, column=EVENT_CONTENT, value=content)

            if type == 'Message':
                total_messages = user_worksheet.cell(row=HEADER_DATA_ROW, column=TOTAL_MESSAGE).value
                user_worksheet.cell(row=HEADER_DATA_ROW, column=TOTAL_MESSAGE, value=total_messages + 1)
                print(f'Total messages for user {user_id} incremented to {total_messages + 1}.')
            elif type == 'Call':
                total_calls = user_worksheet.cell(row=HEADER_DATA_ROW, column=TOTAL_CALL).value
                user_worksheet.cell(row=HEADER_DATA_ROW, column=TOTAL_CALL, value=total_calls + 1)
                print(f'Total calls for user {user_id} incremented to {total_calls + 1}.')
            print(f'Updated sheet {user_id} with contents [{type}, {time}, {content}].')

            print('Saving workbook...')
            workbook.save(WORKBOOK_PATH)
            return

    print(f'No sheet found for user {user_id}. Creating new sheet...')

    createUserSheet(user_id, 'N/A', 'N/A')
    updateSpreadsheet(user_id, type, time, content)


def getLastTimeOfEvent(user_id, event_type):
    """
    Retrieve the last event time of a specific type for a user.

    Args:
    user_id (str): The ID of the user.
    event_type (str): The type of event ('message', 'call', 'join', 'leave').
    """

    print('------')
    print(f'Getting last time of event type \'{event_type}\' for user {user_id}.')

    for id in workbook.sheetnames:
        if id == user_id:
            user_worksheet = workbook[id]
            for row in user_worksheet.iter_rows(min_row=LOG_DATA_ROW):
                if row[0].value == event_type:
                    print(f'Last event {event_type} for user {user_id} found: {row[EVENT_TIME - 1].value}.')
                    return row[EVENT_TIME - 1].value
                
    print(f'Event {event_type} for user {user_id} not found.')
    return 'N/A'


def getTotalMessages(user_id):
    """
    Retrieve the total number of messages sent by a user.

    Args:
    user_id (str): The ID of the user.
    """

    print('------')
    print(f'Getting total messages for user {user_id}...')

    total_messages = 0

    for id in workbook.sheetnames:
        if id == user_id:
            user_worksheet = workbook[id]
            total_messages = user_worksheet.cell(row=HEADER_DATA_ROW, column=TOTAL_MESSAGE).value

    print(f'Total messages for user {user_id}: {total_messages}')

    return total_messages

def getTotalCalls(user_id):
    """
    Retrieve the total number of voice calls joined by a user.

    Args:
    user_id (str): The ID of the user.
    """

    print('------')
    print(f'Getting total calls for user {user_id}.')

    total_calls = 0

    for id in workbook.sheetnames:
        if id == user_id:
            user_worksheet = workbook[id]
            total_calls = user_worksheet.cell(row=HEADER_DATA_ROW, column=TOTAL_CALL).value

    print(f'Total calls for user {user_id}: {total_calls}')

    return total_calls

def getInviter(user_id):
    """
    Retrieve the inviter of a user.

    Args:
    user_id (str): The ID of the user.
    """

    print('------')
    print(f'Getting inviter for user {user_id}.')

    inviter = 'N/A'

    for id in workbook.sheetnames:
        if id == user_id:
            user_worksheet = workbook[id]
            inviter = user_worksheet.cell(row=HEADER_DATA_ROW, column=INVITED_BY).value

    print(f'Inviter for user {user_id}: {inviter}')

    return inviter

# ---------------------------------------------------------------------------
# Discord Bot Setup
# ---------------------------------------------------------------------------
class Client(commands.Bot):
    """Discord bot for tracking user text and voice activity."""
    
    async def on_ready(self):
        """Triggered on bot initialization"""
        
        print('------')
        print(f'Logged in as {self.user}')
        
        '''
        If you wish to quickly sync updates to your discord, this is a template to do so
        '''
        
        try:
            synced = await self.tree.sync(guild=discord.Object(id=GUILD_ID))
            print(f'Synced {len(synced)} command(s) to the guild {GUILD_ID}.')
        except Exception as e:
            print(f'Error syncing commands: {e}')
        

    async def on_message(self, message):
        """Triggered when a user sends a message."""

        # exiting if bot message or empty message
        if message.author.bot:
            return
        elif message.content == (''):
            return

        print('------')
        print(f'Message from {message.author}: {message.content}')

        updateSpreadsheet(message.author.name, 
                          'Message', 
                          datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 
                          message.content)


    async def on_voice_state_update(self, member, before, after):
        """Triggered when a user joins a voice channel."""

        if member.bot:
            return

        if before.channel is None and after.channel is not None:  # User joined a voice channel
            print('------')
            print(f'{member.name} joined voice channel {after.channel.name}.')
            updateSpreadsheet(member.name, 
                              'Call', 
                              datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 
                              'Joined ' + after.channel.name)
            

    async def on_member_join(self, member):
        """Triggered when a new member joins the server."""

        print('------')
        print(f'Member {member.name} has joined the server.')
        
        if member.bot:
            return

        time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        createUserSheet(member.name, 
                          'N/A', 
                          time)
        updateSpreadsheet(member.name,
                          'Join', 
                          time, 
                          'Joined the server')

        message = member.name + ' has joined the server.\n'

        channel = self.guilds[0].system_channel

        await channel.send(message)

    async def on_member_remove(self, member):
        """Triggered when a member leaves the server."""

        print('------')
        print(f'Member {member.name} has left the server.')

        if member.bot:
            return

        updateSpreadsheet(member.name, 
                          'Leave', 
                          datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 
                          'Left the server')
        
        message = member.name + ' has left the server.\n'
        message += 'They were originally invited by ' + getInviter(member.name) + ' on ' + getLastTimeOfEvent(member.name, 'Join') + '.\n'
        message += 'They sent a total of ' + str(getTotalMessages(member.name)) + ' messages.\n'
        message += 'They joined a total of ' + str(getTotalCalls(member.name)) + ' voice calls.\n'

        print('notifying server...')

        channel = self.guilds[0].system_channel

        await channel.send(message)

        print('server notified.')


# ---------------------------------------------------------------------------
# Discord Slash Commands
# ---------------------------------------------------------------------------
intents = discord.Intents.default()
intents.message_content = True
intents.members = True
intents.guilds = True
client = Client(command_prefix='!', intents=intents)

@client.tree.command(name="get-user-activity", description="Get a user's activity summary", guild=discord.Object(id=GUILD_ID))
async def activity_check(interaction: discord.Interaction, user: discord.Member):
    print('------')
    print(f'ActivityCheck command invoked by {interaction.user} for user {user}.')
    print('Checking User Activity and Displaying...')

    lastMessageDateTime = getLastTimeOfEvent(user.name, 'Message')
    lastVoiceJoinDateTime = getLastTimeOfEvent(user.name, 'Call')
    totalMessages = getTotalMessages(user.name)
    totalCalls = getTotalCalls(user.name)

    message = f'User {user.mention} Activity Summary:\n'
    message += '-----------------------------------------------------\n'
    message += f'Last Message Sent:        {lastMessageDateTime}\n'
    message += f'Last Voice Call Joined:   {lastVoiceJoinDateTime}\n'
    message += '-----------------------------------------------------\n'
    message += f'Total Messages Sent:      {totalMessages}\n'
    message += f'Total Voice Calls Joined: {totalCalls}\n'

    await interaction.response.send_message(message)

    print('User Activity Displayed.')

@client.tree.command(name="get-server-activity", description="Check Last Server Activity", guild=discord.Object(id=GUILD_ID))
async def server_activity_check(interaction: discord.Interaction):
    print('------')
    print(f'ServerActivityCheck command invoked by {interaction.user}.')
    print('Checking Server Activity and Displaying...')

    userData = []

    for id in workbook.sheetnames:
        userData.append((id, getTotalMessages(id), getTotalCalls(id)))

    message = 'Most Messages Sent:\n'
    userData.sort(key=lambda x: x[1], reverse=True)
    for user in userData:
        message += f'{user[0]}: {user[1]} messages\n'
    message += '-----------------------------------------------------\n'
    message += 'Most Voice Calls Joined:\n'
    userData.sort(key=lambda x: x[2], reverse=True)
    for user in userData:
        message += f'{user[0]}: {user[2]} calls\n'

    await interaction.response.send_message(message)

    print('Server Activity Displayed.')


# ---------------------------------------------------------------------------
# Entry Point
# ---------------------------------------------------------------------------
client.run(DISCORD_TOKEN)